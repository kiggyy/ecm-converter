import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from bcolors import bcolors

FLOAT_FORMAT = "F"
INT_FORMAT = "I"

#FLOAT_FORMAT_STR = "#,###.??;(#,###.??);0"
#INT_FORMAT_STR = "0"

MAPPING_COLUMNS = {
    "Feeder": {"size":7},
    "Value": {"size":20},
    "Footprint": {"size":10},
    "PartRemark": {"size":14},
    "X": {"size":5, "format" : FLOAT_FORMAT},
    "Y": {"size":5, "format" : FLOAT_FORMAT},
    "H": {"size":5, "format" : FLOAT_FORMAT},
    "W": {"size":4, "format" : INT_FORMAT},
    "Arot": {"size":5, "format" : FLOAT_FORMAT},
    "Nz": {"size":4, "format" : INT_FORMAT},
    "Xofs": {"size":5, "format" : FLOAT_FORMAT},
    "Yofs": {"size":5, "format" : FLOAT_FORMAT},
    "Afed": {"size":4, "format" : INT_FORMAT},
    "Strk": {"size":4, "format" : INT_FORMAT},
    "FIdx": {"size":4, "format" : INT_FORMAT},
    "Designators": {"size":50},
    "PartRemarkTS": {"size":5},
    "PartRemark3P": {"size":1},
}
MAPPING_HEADERS = list(MAPPING_COLUMNS.keys())


class Mapping:
    def __init__(self, mapping_file) -> None:
        self.mapping_file = mapping_file
        self.current_mapping_values = {}
        self.load_mapping()

    def __init_excel_mapping(self, sheet) -> None:
        column_sizes = list(MAPPING_COLUMNS.values())
        for i in range(len(MAPPING_HEADERS)):
            col = i + 1
            sheet.cell(row=1, column=col).value = MAPPING_HEADERS[i]
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = column_sizes[i]["size"]

    def __prepare_current_column_layout(self) -> None:
        self.mapping_headers = {}
        row = self.sheet[1]
        for cell in row:
            self.mapping_headers[cell.value] = cell.col_idx
            
    def __get_current_mapping_values(self) -> None:
        if self.__is_new:
            self.sheet_rows = 1
            return
        col_value = self.__get_column_by_name("Value") - 1
        col_footprint = self.__get_column_by_name("Footprint") - 1
        self.sheet_rows = 0
        for row in self.sheet.rows:
            self.sheet_rows += 1
            if self.sheet_rows > 1:
                if row[col_value].value is None or row[col_footprint].value is None:
                    continue
                key = row[col_value].value + "#:#" + row[col_footprint].value
                self.current_mapping_values[key] = self.sheet_rows

    def load_mapping(self) -> None:
        if os.path.exists(self.mapping_file):
            self.workbook = load_workbook(filename=self.mapping_file)
            self.__is_new = False
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.__is_new = True
            self.__init_excel_mapping(self.sheet)

        self.__is_resolved = False
        self.sheet = self.workbook.active
        self.__prepare_current_column_layout()
        self.__get_current_mapping_values()

    def find_row_by_cell(self, index, value):
        idx = 0
        index = index - 1
        for row in self.sheet.rows:
            idx += 1
            if row[index].value == value:
                return idx
        return 0
    
    def check_repeat_feeder() -> bool:
        pass

    def merge_mapping(self, pcb_new_mapping) -> int:
        if self.__is_new:
            self.changes_count = -1
            return self.changes_count
        self.changes_count = 0
        used_feeders = []
        self.__is_resolved = True
        current_mapping_values = self.current_mapping_values.copy()
        for key, p in pcb_new_mapping.items():
            val = p["Value"]
            if val != val:
                continue
            row_no = self.current_mapping_values.get(key, 0)
            if row_no:
                row = self.sheet[row_no]
#                print("Found {}, row {}".format(val, row_no))
                p["row"] = row_no  # update the pointed cell
                for i in MAPPING_COLUMNS:
                    if i in ["Value","Footprint", "Designators"]:
                        continue
                    p[i] = row[self.__get_column_by_name(i) - 1].value
                    if p[i] is None:
                        if  i in ["Afed", "PartRemark3P"]: #TODO: afed must be initeger
                           p[i] = ""
                        else: 
                            self.__is_resolved = False
                            bcolors.color_print_warning(
                                "Row {}: part {}/{} has zero attribute {}".format(
                                    row_no, val, p["Footprint"], i
                                )
                            )

                # check feeder
                feeder = p["Feeder"]
                if feeder not in used_feeders:
                    used_feeders.append(feeder)
                else:
                    if feeder and feeder < 999 :
                        bcolors.color_print_warning("Feeder {} used multiple times".format(feeder))
                        self.__is_resolved = False
                    
                #update designators
                col = self.__get_column_by_name("Designators") - 1
                
                designators_projects = [] if row[col].value is None else row[col].value.split("\n")
                project_name = p["Designators"].split(': ')[0] + ': '
                is_found = False
                for project_designators in designators_projects:
                    if project_designators.upper().startswith(project_name):
                        is_found = True
                        if p["Designators"] != project_designators:
                            designators_projects[designators_projects.index(project_designators)] = p["Designators"]
                            self.changes_count += 1
                            print(
                                "Row {}: part {}/{} changed {}: {} -> {}".format(
                                    row_no, val, p["Footprint"], i, project_designators, p["Designators"]
                                )
                            )
                            break
                if not is_found:
                    self.changes_count += 1
                    designators_projects.append(p["Designators"])
                    print(
                        "Row {}: part {}/{} new project designators: -> {}".format(
                            row_no, val, p["Footprint"], p["Designators"]
                        )
                    )
                        
                p["Designators"] = "\n".join(designators_projects)

                current_mapping_values[val] = 0
            else:
                self.changes_count += 1
                self.__is_resolved = False
                print("New value {}".format(val))

        # absent = [k for k in current_mapping_values.keys() if current_mapping_values[k] != 0]
        # for p in absent:
        #     self.__is_resolved = False
        #     print("Absent: {}".format(p))

        print("Mapping changes: {}".format(self.changes_count))
        return self.changes_count

    def __get_column_by_name(self, name) -> int:
        if name in self.mapping_headers:
            return self.mapping_headers[name]
        return -1

    def save_mapping(self, mapping) -> None:
        self.sheet = self.workbook.active
        row = 2

        for p in mapping.values():
            row = p['row']
            if row < 1:
                self.sheet_rows+= 1
                row = self.sheet_rows
                p['row'] = row
            for k in p:
                col = self.__get_column_by_name(k)
                if col != -1:
                    cell = self.sheet.cell(row=row, column=col)
                    cell.value = p[k]
                    if k == "Designators":
                        cell.alignment = Alignment(wrapText=True)
                    if p[k] and "format" in MAPPING_COLUMNS[k]:
                        format = MAPPING_COLUMNS[k]["format"]
                        if format == FLOAT_FORMAT:
                            cell.value = float(p[k])
                        elif format == INT_FORMAT:
                            cell.value = int(p[k])
                        #cell.number_format = MAPPING_COLUMNS[k]["format"]

            row += 1
        if not self.__is_new:
            shutil.copy(self.mapping_file, self.mapping_file + ".bak")
        try:
            self.workbook.save(filename=self.mapping_file)
        except Exception as err:
            print("CAN't update mapping file {}: {}\nExiting!".format(self.mapping_file, err))
            exit(100)
            
    def is_resolved(self) -> bool:
        return self.__is_resolved
